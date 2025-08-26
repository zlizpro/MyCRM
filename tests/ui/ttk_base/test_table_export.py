"""
MiniCRM TTK表格导出组件测试

测试TableExportTTK组件的功能，确保导出功能正常工作。
"""

import csv
import os
import tempfile
import tkinter as tk
import unittest
from unittest.mock import Mock, patch

from src.minicrm.ui.ttk_base.table_export_ttk import (
    ExportFormat,
    ExportProgress,
    ExportScope,
    TableExportTTK,
)


try:
    import openpyxl

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class TestExportProgress(unittest.TestCase):
    """测试ExportProgress类"""

    def test_progress_initialization(self):
        """测试进度对象初始化"""
        progress = ExportProgress(100)

        self.assertEqual(progress.total, 100)
        self.assertEqual(progress.current, 0)
        self.assertEqual(progress.message, "")
        self.assertFalse(progress.cancelled)
        self.assertEqual(progress.percentage, 0.0)

    def test_progress_update(self):
        """测试进度更新"""
        progress = ExportProgress(100)

        progress.update(50, "正在处理...")

        self.assertEqual(progress.current, 50)
        self.assertEqual(progress.message, "正在处理...")
        self.assertEqual(progress.percentage, 50.0)

    def test_progress_cancel(self):
        """测试取消进度"""
        progress = ExportProgress(100)

        progress.cancel()

        self.assertTrue(progress.cancelled)

    def test_progress_percentage_edge_cases(self):
        """测试进度百分比边界情况"""
        # 总数为0的情况
        progress = ExportProgress(0)
        self.assertEqual(progress.percentage, 0.0)

        # 超过100%的情况
        progress = ExportProgress(100)
        progress.update(150)
        self.assertEqual(progress.percentage, 100.0)


class TestTableExportTTK(unittest.TestCase):
    """测试TableExportTTK组件"""

    def setUp(self):
        """测试准备"""
        self.root = tk.Tk()
        self.root.withdraw()  # 隐藏测试窗口

        self.columns = [
            {"id": "name", "text": "姓名"},
            {"id": "age", "text": "年龄"},
            {"id": "city", "text": "城市"},
            {"id": "phone", "text": "电话"},
        ]

        self.export_widget = TableExportTTK(
            self.root,
            columns=self.columns,
            enable_excel=EXCEL_AVAILABLE,
            enable_csv=True,
            show_progress=False,  # 禁用进度对话框以便测试
        )

        # 测试数据
        self.test_data = [
            {"name": "张三", "age": "25", "city": "北京", "phone": "13800138000"},
            {"name": "李四", "age": "30", "city": "上海", "phone": "13800138001"},
            {"name": "王五", "age": "35", "city": "广州", "phone": "13800138002"},
            {"name": "赵六", "age": "28", "city": "深圳", "phone": "13800138003"},
            {"name": "小张", "age": "22", "city": "北京", "phone": "13800138004"},
        ]

    def tearDown(self):
        """测试清理"""
        self.export_widget.destroy()
        self.root.destroy()

    def test_export_initialization(self):
        """测试导出组件初始化"""
        self.assertEqual(len(self.export_widget.columns), 4)
        self.assertEqual(self.export_widget.enable_csv, True)
        self.assertEqual(self.export_widget.enable_excel, EXCEL_AVAILABLE)
        self.assertFalse(self.export_widget.show_progress)

    def test_export_to_csv(self):
        """测试CSV导出功能"""
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".csv", delete=False
        ) as temp_file:
            temp_filename = temp_file.name

        try:
            # 执行CSV导出
            success = self.export_widget._export_to_csv(
                temp_filename, self.test_data, self.columns
            )

            self.assertTrue(success)
            self.assertTrue(os.path.exists(temp_filename))

            # 验证CSV内容
            with open(temp_filename, encoding="utf-8-sig") as csvfile:
                reader = csv.reader(csvfile)
                rows = list(reader)

                # 验证标题行
                expected_headers = ["姓名", "年龄", "城市", "电话"]
                self.assertEqual(rows[0], expected_headers)

                # 验证数据行数
                self.assertEqual(len(rows), len(self.test_data) + 1)  # +1 for header

                # 验证第一行数据
                self.assertEqual(rows[1], ["张三", "25", "北京", "13800138000"])

        finally:
            # 清理临时文件
            if os.path.exists(temp_filename):
                os.unlink(temp_filename)

    @unittest.skipUnless(EXCEL_AVAILABLE, "openpyxl not available")
    def test_export_to_excel(self):
        """测试Excel导出功能"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            temp_filename = temp_file.name

        try:
            # 执行Excel导出
            success = self.export_widget._export_to_excel(
                temp_filename, self.test_data, self.columns
            )

            self.assertTrue(success)
            self.assertTrue(os.path.exists(temp_filename))

            # 验证Excel内容
            workbook = openpyxl.load_workbook(temp_filename)
            worksheet = workbook.active

            # 验证标题行
            expected_headers = ["姓名", "年龄", "城市", "电话"]
            actual_headers = [cell.value for cell in worksheet[1]]
            self.assertEqual(actual_headers, expected_headers)

            # 验证数据行数
            self.assertEqual(
                worksheet.max_row, len(self.test_data) + 1
            )  # +1 for header

            # 验证第一行数据
            first_data_row = [cell.value for cell in worksheet[2]]
            self.assertEqual(first_data_row, ["张三", "25", "北京", "13800138000"])

        finally:
            # 清理临时文件
            if os.path.exists(temp_filename):
                os.unlink(temp_filename)

    def test_export_data_directly_csv(self):
        """测试直接导出CSV数据"""
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as temp_file:
            temp_filename = temp_file.name

        try:
            # 直接导出数据
            success = self.export_widget.export_data_directly(
                self.test_data, temp_filename, ExportFormat.CSV
            )

            self.assertTrue(success)
            self.assertTrue(os.path.exists(temp_filename))

            # 验证文件内容
            with open(temp_filename, encoding="utf-8-sig") as csvfile:
                reader = csv.reader(csvfile)
                rows = list(reader)
                self.assertEqual(len(rows), len(self.test_data) + 1)

        finally:
            if os.path.exists(temp_filename):
                os.unlink(temp_filename)

    @unittest.skipUnless(EXCEL_AVAILABLE, "openpyxl not available")
    def test_export_data_directly_excel(self):
        """测试直接导出Excel数据"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            temp_filename = temp_file.name

        try:
            # 直接导出数据
            success = self.export_widget.export_data_directly(
                self.test_data, temp_filename, ExportFormat.EXCEL
            )

            self.assertTrue(success)
            self.assertTrue(os.path.exists(temp_filename))

            # 验证文件内容
            workbook = openpyxl.load_workbook(temp_filename)
            worksheet = workbook.active
            self.assertEqual(worksheet.max_row, len(self.test_data) + 1)

        finally:
            if os.path.exists(temp_filename):
                os.unlink(temp_filename)

    def test_export_empty_data(self):
        """测试导出空数据"""
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as temp_file:
            temp_filename = temp_file.name

        try:
            # 导出空数据
            success = self.export_widget.export_data_directly(
                [], temp_filename, ExportFormat.CSV
            )

            self.assertFalse(success)

        finally:
            if os.path.exists(temp_filename):
                os.unlink(temp_filename)

    def test_export_selected_columns(self):
        """测试导出选定列"""
        # 只导出姓名和年龄列
        selected_columns = [
            {"id": "name", "text": "姓名"},
            {"id": "age", "text": "年龄"},
        ]

        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as temp_file:
            temp_filename = temp_file.name

        try:
            success = self.export_widget.export_data_directly(
                self.test_data, temp_filename, ExportFormat.CSV, selected_columns
            )

            self.assertTrue(success)

            # 验证只有选定的列被导出
            with open(temp_filename, encoding="utf-8-sig") as csvfile:
                reader = csv.reader(csvfile)
                rows = list(reader)

                # 验证标题行只有两列
                self.assertEqual(rows[0], ["姓名", "年龄"])

                # 验证数据行只有两列
                self.assertEqual(rows[1], ["张三", "25"])

        finally:
            if os.path.exists(temp_filename):
                os.unlink(temp_filename)

    def test_export_callbacks(self):
        """测试导出回调函数"""
        start_callback = Mock()
        complete_callback = Mock()
        progress_callback = Mock()

        self.export_widget.on_export_started = start_callback
        self.export_widget.on_export_completed = complete_callback
        self.export_widget.on_export_progress = progress_callback

        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as temp_file:
            temp_filename = temp_file.name

        try:
            # 模拟导出过程
            self.export_widget._on_export_finished(temp_filename, True)

            # 验证完成回调被调用
            complete_callback.assert_called_once_with(temp_filename, True)

        finally:
            if os.path.exists(temp_filename):
                os.unlink(temp_filename)

    def test_export_progress_with_cancellation(self):
        """测试带取消功能的导出进度"""
        progress = ExportProgress(100)
        self.export_widget.export_progress = progress

        # 模拟取消导出
        progress.cancel()

        # 验证导出被取消
        success = self.export_widget._export_to_csv(
            "dummy.csv", self.test_data, self.columns
        )

        # 由于进度被取消，导出应该失败
        self.assertFalse(success)

    def test_unsupported_export_format(self):
        """测试不支持的导出格式"""
        with tempfile.NamedTemporaryFile(suffix=".txt", delete=False) as temp_file:
            temp_filename = temp_file.name

        try:
            # 尝试使用不支持的格式
            success = self.export_widget.export_data_directly(
                self.test_data,
                temp_filename,
                "unsupported_format",  # 这会导致错误
            )

            self.assertFalse(success)

        finally:
            if os.path.exists(temp_filename):
                os.unlink(temp_filename)

    def test_cleanup(self):
        """测试资源清理"""
        # 设置一些状态
        self.export_widget.export_progress = ExportProgress(100)
        self.export_widget.on_export_started = Mock()
        self.export_widget.on_export_completed = Mock()

        # 执行清理
        self.export_widget.cleanup()

        # 验证状态被清理
        self.assertIsNone(self.export_widget.on_export_started)
        self.assertIsNone(self.export_widget.on_export_completed)

    @patch("tkinter.filedialog.asksaveasfilename")
    def test_show_export_dialog_cancel(self, mock_filedialog):
        """测试导出对话框取消操作"""
        # 模拟用户取消文件选择
        mock_filedialog.return_value = ""

        # 这个测试主要验证不会抛出异常
        # 实际的对话框测试需要更复杂的UI测试框架
        try:
            # 创建一个简单的测试对话框
            dialog = tk.Toplevel(self.root)
            # 创建模拟的列变量字典
            column_vars = {}
            for col in self.columns:
                column_vars[col["id"]] = tk.BooleanVar(value=True)

            self.export_widget._start_export(
                dialog,
                ExportFormat.CSV.value,
                ExportScope.ALL_DATA.value,
                self.test_data,
                None,
                None,
                column_vars,
            )
            dialog.destroy()
        except Exception as e:
            self.fail(f"导出对话框处理失败: {e}")


if __name__ == "__main__":
    unittest.main()
