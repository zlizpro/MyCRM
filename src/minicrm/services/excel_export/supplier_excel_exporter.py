"""
供应商Excel导出器

专门负责供应商数据的Excel导出功能.
支持供应商质量评估和合作分析报表.
"""

import csv
import logging
from typing import Any

from minicrm.core.exceptions import ServiceError

from .excel_formatters import ExcelFormatters


class SupplierExcelExporter:
    """
    供应商Excel导出器

    专门负责供应商数据的Excel导出功能.
    """

    def __init__(self):
        """初始化供应商Excel导出器"""
        self._logger = logging.getLogger(__name__)
        self._formatters = ExcelFormatters()

    def export_supplier_data(
        self, suppliers: list[dict[str, Any]], output_path: str
    ) -> bool:
        """
        导出供应商数据到Excel

        Args:
            suppliers: 供应商数据列表
            output_path: 输出文件路径

        Returns:
            bool: 导出是否成功
        """
        try:
            self._logger.info(f"开始导出供应商数据到Excel: {output_path}")

            # 尝试使用openpyxl库
            try:
                return self._export_with_openpyxl(suppliers, output_path)
            except ImportError:
                # 备用方案:CSV格式
                return self._export_as_csv(
                    suppliers, output_path.replace(".xlsx", ".csv")
                )

        except Exception as e:
            self._logger.error(f"导出供应商数据失败: {e}")
            raise ServiceError(
                f"导出供应商数据失败: {e}", "SupplierExcelExporter"
            ) from e

    def _export_with_openpyxl(
        self, suppliers: list[dict[str, Any]], output_path: str
    ) -> bool:
        """使用openpyxl导出供应商数据"""
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "供应商数据"

            # 设置列标题
            headers = [
                "供应商ID",
                "供应商名称",
                "联系人",
                "电话",
                "邮箱",
                "地址",
                "供应商类别",
                "质量评分",
                "交付评分",
                "价格竞争力",
                "合作年限",
                "状态",
            ]

            # 写入标题行
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                # 使用绿色主题区分供应商数据
                self._formatters.apply_openpyxl_style(cell, "header")
                # 自定义供应商标题颜色
                from openpyxl.styles import PatternFill

                cell.fill = PatternFill(
                    start_color="70AD47", end_color="70AD47", fill_type="solid"
                )

            # 写入数据行
            for row, supplier in enumerate(suppliers, 2):
                data = [
                    supplier.get("id", ""),
                    supplier.get("name", ""),
                    supplier.get("contact_person", ""),
                    supplier.get("phone", ""),
                    supplier.get("email", ""),
                    supplier.get("address", ""),
                    supplier.get("category", ""),
                    f"{supplier.get('quality_score', 0):.1f}",
                    f"{supplier.get('delivery_score', 0):.1f}",
                    supplier.get("price_competitive", ""),
                    f"{supplier.get('cooperation_years', 0)}年",
                    supplier.get("status", ""),
                ]

                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    # 数字列使用数字格式
                    style = "number" if col in [8, 9] else "data"
                    self._formatters.apply_openpyxl_style(cell, style)

            # 自动调整列宽
            self._formatters.auto_adjust_column_width(ws, len(headers))

            wb.save(output_path)
            wb.close()

            self._logger.info(f"导出供应商数据成功: {output_path}")
            return True

        except ImportError:
            raise  # 重新抛出ImportError以便使用备用方案
        except Exception as e:
            self._logger.error(f"使用openpyxl导出供应商数据失败: {e}")
            return False

    def _export_as_csv(self, suppliers: list[dict[str, Any]], output_path: str) -> bool:
        """导出供应商数据为CSV格式"""
        try:
            with open(output_path, "w", newline="", encoding="utf-8-sig") as csvfile:
                fieldnames = [
                    "供应商ID",
                    "供应商名称",
                    "联系人",
                    "电话",
                    "邮箱",
                    "地址",
                    "供应商类别",
                    "质量评分",
                    "交付评分",
                    "价格竞争力",
                    "合作年限",
                    "状态",
                ]

                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()

                for supplier in suppliers:
                    row = {
                        "供应商ID": supplier.get("id", ""),
                        "供应商名称": supplier.get("name", ""),
                        "联系人": supplier.get("contact_person", ""),
                        "电话": supplier.get("phone", ""),
                        "邮箱": supplier.get("email", ""),
                        "地址": supplier.get("address", ""),
                        "供应商类别": supplier.get("category", ""),
                        "质量评分": f"{supplier.get('quality_score', 0):.1f}",
                        "交付评分": f"{supplier.get('delivery_score', 0):.1f}",
                        "价格竞争力": supplier.get("price_competitive", ""),
                        "合作年限": f"{supplier.get('cooperation_years', 0)}年",
                        "状态": supplier.get("status", ""),
                    }
                    writer.writerow(row)

            self._logger.info(f"导出供应商数据为CSV成功: {output_path}")
            return True

        except Exception as e:
            self._logger.error(f"导出供应商数据为CSV失败: {e}")
            return False
