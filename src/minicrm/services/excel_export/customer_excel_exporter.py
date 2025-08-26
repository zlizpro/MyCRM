"""
客户Excel导出器

专门负责客户数据的Excel导出功能.
支持多种格式和详细的客户分析报表.
"""

import csv
import logging
from datetime import datetime
from typing import Any

from minicrm.core.exceptions import ServiceError

from .excel_formatters import ExcelFormatters
from .excel_statistics_calculator import ExcelStatisticsCalculator


class CustomerExcelExporter:
    """
    客户Excel导出器

    专门负责客户数据的Excel导出功能.
    """

    def __init__(self):
        """初始化客户Excel导出器"""
        self._logger = logging.getLogger(__name__)
        self._formatters = ExcelFormatters()
        self._calculator = ExcelStatisticsCalculator()

    def export_customer_data(
        self,
        customers: list[dict[str, Any]],
        output_path: str,
        include_analysis: bool = True,
    ) -> bool:
        """
        导出客户数据到Excel

        Args:
            customers: 客户数据列表
            output_path: 输出文件路径
            include_analysis: 是否包含分析数据

        Returns:
            bool: 导出是否成功
        """
        try:
            self._logger.info(f"开始导出客户数据到Excel: {output_path}")

            # 尝试使用openpyxl库
            try:
                return self._export_with_openpyxl(
                    customers, output_path, include_analysis
                )
            except ImportError:
                # 备用方案:使用xlsxwriter
                return self._export_with_xlsxwriter(
                    customers, output_path, include_analysis
                )

        except Exception as e:
            self._logger.error(f"导出客户数据失败: {e}")
            raise ServiceError(f"导出客户数据失败: {e}", "CustomerExcelExporter") from e

    def _export_with_openpyxl(
        self, customers: list[dict[str, Any]], output_path: str, include_analysis: bool
    ) -> bool:
        """使用openpyxl导出客户数据"""
        try:
            from openpyxl import Workbook

            wb = Workbook()

            try:
                # 删除默认工作表
                wb.remove(wb.active)

                # 创建客户基本信息工作表
                ws_basic = wb.create_sheet("客户基本信息")
                self._create_basic_sheet_openpyxl(ws_basic, customers)

                # 创建客户分析工作表
                if include_analysis:
                    ws_analysis = wb.create_sheet("客户分析")
                    self._create_analysis_sheet_openpyxl(ws_analysis, customers)

                    # 创建客户价值分布工作表
                    ws_value = wb.create_sheet("价值分布")
                    self._create_value_sheet_openpyxl(ws_value, customers)

                # 创建汇总工作表
                ws_summary = wb.create_sheet("数据汇总")
                self._create_summary_sheet_openpyxl(ws_summary, customers)

                # 保存文件
                wb.save(output_path)

            finally:
                # 确保工作簿被正确关闭
                wb.close()

            self._logger.info(f"使用openpyxl导出客户数据成功: {output_path}")
            return True

        except ImportError:
            raise  # 重新抛出ImportError以便使用备用方案
        except PermissionError as e:
            self._logger.error(f"权限不足,无法写入文件 {output_path}: {e}")
            return False
        except OSError as e:
            self._logger.error(f"文件系统错误,无法保存文件 {output_path}: {e}")
            return False
        except Exception as e:
            self._logger.error(f"使用openpyxl导出客户数据失败: {e}")
            return False

    def _create_basic_sheet_openpyxl(
        self, ws: Any, customers: list[dict[str, Any]]
    ) -> None:
        """创建客户基本信息工作表"""
        # 设置列标题
        headers = [
            "客户ID",
            "客户名称",
            "联系人",
            "电话",
            "邮箱",
            "地址",
            "行业",
            "公司规模",
            "客户类型",
            "创建日期",
            "最后互动",
            "状态",
        ]

        # 写入标题行
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._formatters.apply_openpyxl_style(cell, "header")

        # 写入数据行
        for row, customer in enumerate(customers, 2):
            data = [
                customer.get("id", ""),
                customer.get("name", ""),
                customer.get("contact_person", ""),
                customer.get("phone", ""),
                customer.get("email", ""),
                customer.get("address", ""),
                customer.get("industry", ""),
                customer.get("company_size", ""),
                customer.get("customer_type", ""),
                customer.get("created_at", ""),
                customer.get("last_interaction_date", ""),
                customer.get("status", ""),
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                # 日期列使用特殊格式
                style = "date" if col in [10, 11] else "data"
                self._formatters.apply_openpyxl_style(cell, style)

        # 自动调整列宽和冻结首行
        self._formatters.auto_adjust_column_width(ws, len(headers))
        self._formatters.freeze_header_row(ws)

    def _create_analysis_sheet_openpyxl(
        self, ws: Any, customers: list[dict[str, Any]]
    ) -> None:
        """创建客户分析工作表"""
        # 分析标题
        title_cell = ws.cell(row=1, column=1, value="客户分析报告")
        self._formatters.apply_openpyxl_style(title_cell, "header")

        ws.cell(
            row=2,
            column=1,
            value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        )

        # 基本统计
        stats_title = ws.cell(row=4, column=1, value="基本统计")
        self._formatters.apply_openpyxl_style(stats_title, "subheader")

        stats = self._calculator.calculate_customer_statistics(customers)
        stats_data = [
            ["客户总数", len(customers)],
            ["活跃客户", stats.get("active_customers", 0)],
            ["高价值客户", stats.get("high_value_customers", 0)],
            ["本月新增", stats.get("new_this_month", 0)],
            ["平均合作时长", f"{stats.get('avg_cooperation_months', 0):.1f}个月"],
        ]

        for i, (label, value) in enumerate(stats_data, 5):
            label_cell = ws.cell(row=i, column=1, value=label)
            value_cell = ws.cell(row=i, column=2, value=value)
            self._formatters.apply_openpyxl_style(label_cell, "data")
            self._formatters.apply_openpyxl_style(value_cell, "data")

        # 行业分布
        industry_title = ws.cell(row=11, column=1, value="行业分布")
        self._formatters.apply_openpyxl_style(industry_title, "subheader")

        industry_dist = self._calculator.calculate_industry_distribution(customers)

        # 行业分布表头
        headers = ["行业", "客户数", "占比"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=12, column=col, value=header)
            self._formatters.apply_openpyxl_style(cell, "subheader")

        total_customers = len(customers)
        for i, (industry, count) in enumerate(industry_dist.items(), 13):
            ws.cell(row=i, column=1, value=industry)
            ws.cell(row=i, column=2, value=count)
            percentage = (
                f"{count / total_customers * 100:.1f}%" if total_customers > 0 else "0%"
            )
            ws.cell(row=i, column=3, value=percentage)

    def _create_value_sheet_openpyxl(
        self, ws: Any, customers: list[dict[str, Any]]
    ) -> None:
        """创建客户价值分布工作表"""
        # 价值分布标题
        title_cell = ws.cell(row=1, column=1, value="客户价值分布分析")
        self._formatters.apply_openpyxl_style(title_cell, "header")

        # 计算价值分布
        value_distribution = self._calculator.calculate_value_distribution(customers)

        # 价值分布表
        headers = ["价值等级", "客户数量", "占比", "描述"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self._formatters.apply_openpyxl_style(cell, "header")

        value_descriptions = {
            "高价值": "核心客户,贡献主要收入",
            "中价值": "重要客户,具有增长潜力",
            "低价值": "一般客户,需要培育",
            "潜在": "新客户或待开发客户",
        }

        total_customers = sum(value_distribution.values())
        for i, (level, count) in enumerate(value_distribution.items(), 4):
            ws.cell(row=i, column=1, value=level)
            ws.cell(row=i, column=2, value=count)
            percentage = (
                f"{count / total_customers * 100:.1f}%" if total_customers > 0 else "0%"
            )
            ws.cell(row=i, column=3, value=percentage)
            ws.cell(row=i, column=4, value=value_descriptions.get(level, ""))

        # 顶级客户列表
        top_title = ws.cell(row=9, column=1, value="顶级客户列表")
        self._formatters.apply_openpyxl_style(top_title, "subheader")

        top_customers = self._calculator.get_top_customers(customers)
        top_headers = ["排名", "客户名称", "价值评分", "合作时长", "年收入"]
        for col, header in enumerate(top_headers, 1):
            cell = ws.cell(row=10, column=col, value=header)
            self._formatters.apply_openpyxl_style(cell, "subheader")

        for i, customer in enumerate(top_customers, 11):
            ws.cell(row=i, column=1, value=i - 10)
            ws.cell(row=i, column=2, value=customer.get("name", ""))
            ws.cell(row=i, column=3, value=f"{customer.get('value_score', 0):.1f}")
            ws.cell(
                row=i, column=4, value=f"{customer.get('cooperation_months', 0)}个月"
            )
            ws.cell(row=i, column=5, value=f"¥{customer.get('annual_revenue', 0):,.0f}")

    def _create_summary_sheet_openpyxl(
        self, ws: Any, customers: list[dict[str, Any]]
    ) -> None:
        """创建客户数据汇总工作表"""
        # 汇总标题
        title_cell = ws.cell(row=1, column=1, value="客户数据汇总")
        self._formatters.apply_openpyxl_style(title_cell, "header")

        ws.cell(
            row=2,
            column=1,
            value=f"数据导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        )

        # 总体概况
        overview_title = ws.cell(row=4, column=1, value="总体概况")
        self._formatters.apply_openpyxl_style(overview_title, "subheader")

        summary = self._calculator.calculate_financial_summary(customers)
        summary_data = [
            ["客户总数", f"{summary.get('total_customers', 0):,}"],
            ["总年收入", f"¥{summary.get('total_revenue', 0):,.0f}"],
            ["平均年收入", f"¥{summary.get('avg_revenue', 0):,.0f}"],
            ["数据完整性", f"{summary.get('data_completeness', 0):.1f}%"],
        ]

        for i, (label, value) in enumerate(summary_data, 5):
            label_cell = ws.cell(row=i, column=1, value=label)
            value_cell = ws.cell(row=i, column=2, value=value)
            self._formatters.apply_openpyxl_style(label_cell, "data")
            self._formatters.apply_openpyxl_style(value_cell, "data")

        # 数据质量报告
        quality_title = ws.cell(row=10, column=1, value="数据质量报告")
        self._formatters.apply_openpyxl_style(quality_title, "subheader")

        quality_report = self._calculator.generate_data_quality_report(customers)
        for i, (metric, value) in enumerate(quality_report.items(), 11):
            metric_cell = ws.cell(row=i, column=1, value=metric)
            value_cell = ws.cell(row=i, column=2, value=value)
            self._formatters.apply_openpyxl_style(metric_cell, "data")
            self._formatters.apply_openpyxl_style(value_cell, "data")

    def _export_with_xlsxwriter(
        self, customers: list[dict[str, Any]], output_path: str, include_analysis: bool
    ) -> bool:
        """使用xlsxwriter导出客户数据(备用方案)"""
        try:
            import xlsxwriter

            workbook = xlsxwriter.Workbook(output_path)

            # 创建格式
            header_format = self._formatters.create_xlsxwriter_format(
                workbook, "header"
            )
            data_format = self._formatters.create_xlsxwriter_format(workbook, "data")

            # 创建客户基本信息工作表
            worksheet = workbook.add_worksheet("客户基本信息")
            self._create_basic_sheet_xlsxwriter(
                worksheet, customers, header_format, data_format
            )

            if include_analysis:
                # 创建分析工作表
                analysis_sheet = workbook.add_worksheet("客户分析")
                self._create_analysis_sheet_xlsxwriter(
                    analysis_sheet, customers, workbook
                )

            workbook.close()

            self._logger.info(f"使用xlsxwriter导出客户数据成功: {output_path}")
            return True

        except ImportError:
            # 如果xlsxwriter也没有,使用CSV格式
            return self._export_as_csv(customers, output_path.replace(".xlsx", ".csv"))
        except Exception as e:
            self._logger.error(f"使用xlsxwriter导出客户数据失败: {e}")
            return False

    def _create_basic_sheet_xlsxwriter(
        self,
        worksheet: Any,
        customers: list[dict[str, Any]],
        header_format: Any,
        data_format: Any,
    ) -> None:
        """使用xlsxwriter创建客户基本信息工作表"""
        headers = [
            "客户ID",
            "客户名称",
            "联系人",
            "电话",
            "邮箱",
            "地址",
            "行业",
            "公司规模",
            "客户类型",
            "创建日期",
            "最后互动",
            "状态",
        ]

        # 写入标题
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)

        # 写入数据
        for row, customer in enumerate(customers, 1):
            data = [
                customer.get("id", ""),
                customer.get("name", ""),
                customer.get("contact_person", ""),
                customer.get("phone", ""),
                customer.get("email", ""),
                customer.get("address", ""),
                customer.get("industry", ""),
                customer.get("company_size", ""),
                customer.get("customer_type", ""),
                customer.get("created_at", ""),
                customer.get("last_interaction_date", ""),
                customer.get("status", ""),
            ]

            for col, value in enumerate(data):
                worksheet.write(row, col, value, data_format)

        # 设置列宽
        self._formatters.auto_adjust_column_width(worksheet, len(headers), "xlsxwriter")

    def _create_analysis_sheet_xlsxwriter(
        self, worksheet: Any, customers: list[dict[str, Any]], workbook: Any
    ) -> None:
        """使用xlsxwriter创建客户分析工作表"""
        title_format = self._formatters.create_xlsxwriter_format(workbook, "header")

        worksheet.write(0, 0, "客户分析报告", title_format)
        worksheet.write(
            1, 0, f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        )

        # 基本统计
        stats = self._calculator.calculate_customer_statistics(customers)
        subheader_format = self._formatters.create_xlsxwriter_format(
            workbook, "subheader"
        )

        worksheet.write(3, 0, "基本统计", subheader_format)

        stats_data = [
            ["客户总数", len(customers)],
            ["活跃客户", stats.get("active_customers", 0)],
            ["高价值客户", stats.get("high_value_customers", 0)],
        ]

        data_format = self._formatters.create_xlsxwriter_format(workbook, "data")
        for i, (label, value) in enumerate(stats_data, 4):
            worksheet.write(i, 0, label, data_format)
            worksheet.write(i, 1, value, data_format)

    def _export_as_csv(self, customers: list[dict[str, Any]], output_path: str) -> bool:
        """导出为CSV格式(最后备用方案)"""
        try:
            with open(output_path, "w", newline="", encoding="utf-8-sig") as csvfile:
                if not customers:
                    return True

                fieldnames = [
                    "客户ID",
                    "客户名称",
                    "联系人",
                    "电话",
                    "邮箱",
                    "地址",
                    "行业",
                    "公司规模",
                    "客户类型",
                    "创建日期",
                    "最后互动",
                    "状态",
                ]

                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()

                for customer in customers:
                    row = {
                        "客户ID": customer.get("id", ""),
                        "客户名称": customer.get("name", ""),
                        "联系人": customer.get("contact_person", ""),
                        "电话": customer.get("phone", ""),
                        "邮箱": customer.get("email", ""),
                        "地址": customer.get("address", ""),
                        "行业": customer.get("industry", ""),
                        "公司规模": customer.get("company_size", ""),
                        "客户类型": customer.get("customer_type", ""),
                        "创建日期": customer.get("created_at", ""),
                        "最后互动": customer.get("last_interaction_date", ""),
                        "状态": customer.get("status", ""),
                    }
                    writer.writerow(row)

            self._logger.info(f"导出客户数据为CSV成功: {output_path}")
            return True

        except Exception as e:
            self._logger.error(f"导出客户数据为CSV失败: {e}")
            return False
