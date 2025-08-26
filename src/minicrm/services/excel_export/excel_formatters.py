"""
Excel格式化器

负责Excel文档的样式配置和格式化功能.
提供统一的样式管理和格式化方法.
"""

import logging
from typing import Any


class ExcelFormatters:
    """
    Excel格式化器

    提供Excel文档的样式配置和格式化功能.
    """

    def __init__(self):
        """初始化Excel格式化器"""
        self._logger = logging.getLogger(__name__)
        self._styles = self._initialize_styles()

    def _initialize_styles(self) -> dict[str, dict[str, Any]]:
        """
        初始化Excel样式配置

        Returns:
            Dict[str, Dict[str, Any]]: 样式配置字典
        """
        return {
            "header": {
                "font_bold": True,
                "font_size": 12,
                "font_color": "FFFFFF",
                "bg_color": "2E75B6",
                "border": True,
                "align": "center",
            },
            "subheader": {
                "font_bold": True,
                "font_size": 11,
                "font_color": "333333",
                "bg_color": "E7F3FF",
                "border": True,
                "align": "center",
            },
            "data": {
                "font_size": 10,
                "font_color": "333333",
                "border": True,
                "align": "left",
            },
            "number": {
                "font_size": 10,
                "font_color": "333333",
                "border": True,
                "align": "right",
                "number_format": "#,##0.00",
            },
            "date": {
                "font_size": 10,
                "font_color": "333333",
                "border": True,
                "align": "center",
                "number_format": "yyyy-mm-dd",
            },
            "highlight": {
                "font_bold": True,
                "font_color": "D63384",
                "bg_color": "FFF3F3",
                "border": True,
            },
        }

    def get_style(self, style_name: str) -> dict[str, Any]:
        """
        获取指定样式配置

        Args:
            style_name: 样式名称

        Returns:
            Dict[str, Any]: 样式配置
        """
        return self._styles.get(style_name, self._styles["data"])

    def apply_openpyxl_style(self, cell, style_name: str):
        """
        应用openpyxl样式到单元格

        Args:
            cell: openpyxl单元格对象
            style_name: 样式名称
        """
        try:
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

            style = self.get_style(style_name)

            # 设置字体
            if (
                style.get("font_bold")
                or style.get("font_size")
                or style.get("font_color")
            ):
                cell.font = Font(
                    bold=style.get("font_bold", False),
                    size=style.get("font_size", 10),
                    color=style.get("font_color", "000000"),
                )

            # 设置背景色
            if style.get("bg_color"):
                cell.fill = PatternFill(
                    start_color=style["bg_color"],
                    end_color=style["bg_color"],
                    fill_type="solid",
                )

            # 设置对齐
            if style.get("align"):
                cell.alignment = Alignment(horizontal=style["align"], vertical="center")

            # 设置边框
            if style.get("border"):
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

        except ImportError:
            self._logger.warning("openpyxl库未安装,跳过样式应用")
        except Exception as e:
            self._logger.error(f"应用openpyxl样式失败: {e}")

    def create_xlsxwriter_format(self, workbook, style_name: str):
        """
        创建xlsxwriter格式对象

        Args:
            workbook: xlsxwriter工作簿对象
            style_name: 样式名称

        Returns:
            xlsxwriter格式对象
        """
        try:
            style = self.get_style(style_name)

            format_dict = {}

            if style.get("font_bold"):
                format_dict["bold"] = True
            if style.get("font_size"):
                format_dict["font_size"] = style["font_size"]
            if style.get("font_color"):
                format_dict["font_color"] = f"#{style['font_color']}"
            if style.get("bg_color"):
                format_dict["bg_color"] = f"#{style['bg_color']}"
            if style.get("border"):
                format_dict["border"] = 1
            if style.get("align"):
                format_dict["align"] = style["align"]
                format_dict["valign"] = "vcenter"
            if style.get("number_format"):
                format_dict["num_format"] = style["number_format"]

            return workbook.add_format(format_dict)

        except Exception as e:
            self._logger.error(f"创建xlsxwriter格式失败: {e}")
            return workbook.add_format({})

    def auto_adjust_column_width(
        self, worksheet, max_col: int, library: str = "openpyxl"
    ):
        """
        自动调整列宽

        Args:
            worksheet: 工作表对象
            max_col: 最大列数
            library: 使用的库(openpyxl或xlsxwriter)
        """
        try:
            if library == "openpyxl":
                from openpyxl.utils import get_column_letter

                for col in range(1, max_col + 1):
                    worksheet.column_dimensions[get_column_letter(col)].width = 15
            elif library == "xlsxwriter":
                for col in range(max_col):
                    worksheet.set_column(col, col, 15)

        except Exception as e:
            self._logger.error(f"自动调整列宽失败: {e}")

    def freeze_header_row(self, worksheet, library: str = "openpyxl"):
        """
        冻结标题行

        Args:
            worksheet: 工作表对象
            library: 使用的库(openpyxl或xlsxwriter)
        """
        try:
            if library == "openpyxl":
                worksheet.freeze_panes = "A2"
            elif library == "xlsxwriter":
                worksheet.freeze_panes(1, 0)

        except Exception as e:
            self._logger.error(f"冻结标题行失败: {e}")
