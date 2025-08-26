"""MiniCRM TTK表格导出组件

提供完整的表格数据导出功能,支持Excel和CSV格式导出.
支持选择性导出、格式自定义、进度显示和错误处理.

设计特点:
- 支持Excel和CSV格式导出
- 选择性导出(当前页/全部/选中)
- 异步导出避免UI阻塞
- 进度显示和取消功能
- 完整的错误处理机制
"""

import csv
from datetime import datetime
from enum import Enum
import logging
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from typing import Any, Callable, Dict, List, Optional


try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from minicrm.ui.ttk_base.base_widget import BaseWidget


class ExportFormat(Enum):
    """导出格式枚举"""

    CSV = "csv"
    EXCEL = "xlsx"


class ExportScope(Enum):
    """导出范围枚举"""

    CURRENT_PAGE = "current_page"
    ALL_DATA = "all_data"
    SELECTED_ROWS = "selected_rows"


class ExportProgress:
    """导出进度信息"""

    def __init__(self, total: int = 0):
        self.total = total
        self.current = 0
        self.message = ""
        self.cancelled = False

    def update(self, current: int, message: str = ""):
        """更新进度"""
        self.current = current
        self.message = message

    def cancel(self):
        """取消导出"""
        self.cancelled = True

    @property
    def percentage(self) -> float:
        """获取进度百分比"""
        if self.total == 0:
            return 0.0
        return min(100.0, (self.current / self.total) * 100.0)


class TableExportTTK(BaseWidget):
    """TTK表格导出组件

    提供完整的表格数据导出功能,支持多种格式和导出选项.
    可以独立使用,也可以集成到DataTableTTK中.
    """

    def __init__(
        self,
        parent,
        columns: List[Dict[str, Any]],
        enable_excel: bool = EXCEL_AVAILABLE,
        enable_csv: bool = True,
        show_progress: bool = True,
        **kwargs,
    ):
        """初始化导出组件

        Args:
            parent: 父组件
            columns: 列定义列表
            enable_excel: 是否启用Excel导出
            enable_csv: 是否启用CSV导出
            show_progress: 是否显示进度对话框
        """
        # 导出配置
        self.columns = columns
        self.enable_excel = enable_excel and EXCEL_AVAILABLE
        self.enable_csv = enable_csv
        self.show_progress = show_progress

        # 导出状态
        self.current_export_thread = None
        self.export_progress = None
        self.progress_dialog = None

        # 事件回调
        self.on_export_started: Optional[Callable[[str, ExportScope], None]] = None
        self.on_export_completed: Optional[Callable[[str, bool], None]] = None
        self.on_export_progress: Optional[Callable[[ExportProgress], None]] = None

        # 日志记录
        self.logger = logging.getLogger(__name__)

        super().__init__(parent, **kwargs)

    def _setup_ui(self) -> None:
        """设置UI布局"""
        # 导出组件通常不需要独立的UI,而是通过方法调用使用

    def _bind_events(self) -> None:
        """绑定事件"""

    def show_export_dialog(
        self,
        data: List[Dict[str, Any]],
        selected_data: Optional[List[Dict[str, Any]]] = None,
        current_page_data: Optional[List[Dict[str, Any]]] = None,
    ) -> None:
        """显示导出对话框

        Args:
            data: 全部数据
            selected_data: 选中的数据
            current_page_data: 当前页数据
        """
        dialog = tk.Toplevel(self)
        dialog.title("导出数据")
        dialog.geometry("400x300")
        dialog.resizable(False, False)
        dialog.transient(self)
        dialog.grab_set()

        # 居中显示
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")

        # 创建对话框内容
        self._create_export_dialog_content(
            dialog, data, selected_data, current_page_data
        )

    def _create_export_dialog_content(
        self,
        dialog: tk.Toplevel,
        data: List[Dict[str, Any]],
        selected_data: Optional[List[Dict[str, Any]]],
        current_page_data: Optional[List[Dict[str, Any]]],
    ) -> None:
        """创建导出对话框内容"""
        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 导出格式选择
        format_frame = ttk.LabelFrame(main_frame, text="导出格式", padding=10)
        format_frame.pack(fill=tk.X, pady=(0, 10))

        format_var = tk.StringVar(
            value=ExportFormat.EXCEL.value
            if self.enable_excel
            else ExportFormat.CSV.value
        )

        if self.enable_excel:
            ttk.Radiobutton(
                format_frame,
                text="Excel文件 (.xlsx)",
                variable=format_var,
                value=ExportFormat.EXCEL.value,
            ).pack(anchor=tk.W)

        if self.enable_csv:
            ttk.Radiobutton(
                format_frame,
                text="CSV文件 (.csv)",
                variable=format_var,
                value=ExportFormat.CSV.value,
            ).pack(anchor=tk.W)

        # 导出范围选择
        scope_frame = ttk.LabelFrame(main_frame, text="导出范围", padding=10)
        scope_frame.pack(fill=tk.X, pady=(0, 10))

        scope_var = tk.StringVar(value=ExportScope.ALL_DATA.value)

        # 全部数据
        ttk.Radiobutton(
            scope_frame,
            text=f"全部数据 ({len(data)} 条记录)",
            variable=scope_var,
            value=ExportScope.ALL_DATA.value,
        ).pack(anchor=tk.W)

        # 当前页数据
        if current_page_data:
            ttk.Radiobutton(
                scope_frame,
                text=f"当前页数据 ({len(current_page_data)} 条记录)",
                variable=scope_var,
                value=ExportScope.CURRENT_PAGE.value,
            ).pack(anchor=tk.W)

        # 选中数据
        if selected_data:
            ttk.Radiobutton(
                scope_frame,
                text=f"选中数据 ({len(selected_data)} 条记录)",
                variable=scope_var,
                value=ExportScope.SELECTED_ROWS.value,
            ).pack(anchor=tk.W)

        # 列选择
        columns_frame = ttk.LabelFrame(main_frame, text="导出列", padding=10)
        columns_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        # 创建滚动区域
        canvas = tk.Canvas(columns_frame, height=100)
        scrollbar = ttk.Scrollbar(
            columns_frame, orient="vertical", command=canvas.yview
        )
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # 列选择复选框
        column_vars = {}

        # 全选/全不选
        select_all_var = tk.BooleanVar(value=True)
        select_all_cb = ttk.Checkbutton(
            scrollable_frame,
            text="全选",
            variable=select_all_var,
            command=lambda: self._toggle_all_columns(column_vars, select_all_var.get()),
        )
        select_all_cb.pack(anchor=tk.W, pady=(0, 5))

        ttk.Separator(scrollable_frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=5)

        # 各列复选框
        for col in self.columns:
            var = tk.BooleanVar(value=True)
            column_vars[col["id"]] = var

            ttk.Checkbutton(
                scrollable_frame, text=col.get("text", col["id"]), variable=var
            ).pack(anchor=tk.W)

        # 按钮区域
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X)

        ttk.Button(button_frame, text="取消", command=dialog.destroy).pack(
            side=tk.RIGHT, padx=(5, 0)
        )

        ttk.Button(
            button_frame,
            text="导出",
            command=lambda: self._start_export(
                dialog,
                format_var.get(),
                scope_var.get(),
                data,
                selected_data,
                current_page_data,
                column_vars,
            ),
        ).pack(side=tk.RIGHT)

    def _toggle_all_columns(
        self, column_vars: Dict[str, tk.BooleanVar], select_all: bool
    ) -> None:
        """切换所有列的选择状态"""
        for var in column_vars.values():
            var.set(select_all)

    def _start_export(
        self,
        dialog: tk.Toplevel,
        export_format: str,
        export_scope: str,
        data: List[Dict[str, Any]],
        selected_data: Optional[List[Dict[str, Any]]],
        current_page_data: Optional[List[Dict[str, Any]]],
        column_vars: Dict[str, tk.BooleanVar],
    ) -> None:
        """开始导出"""
        # 获取选中的列
        selected_columns = [col for col in self.columns if column_vars[col["id"]].get()]

        if not selected_columns:
            messagebox.showwarning("导出警告", "请至少选择一列进行导出")
            return

        # 确定导出数据
        if export_scope == ExportScope.ALL_DATA.value:
            export_data = data
        elif export_scope == ExportScope.CURRENT_PAGE.value:
            export_data = current_page_data or []
        elif export_scope == ExportScope.SELECTED_ROWS.value:
            export_data = selected_data or []
        else:
            export_data = data

        if not export_data:
            messagebox.showwarning("导出警告", "没有数据可以导出")
            return

        # 选择保存文件
        file_extension = f".{export_format}"
        file_types = []

        if export_format == ExportFormat.EXCEL.value:
            file_types = [("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        elif export_format == ExportFormat.CSV.value:
            file_types = [("CSV文件", "*.csv"), ("所有文件", "*.*")]

        filename = filedialog.asksaveasfilename(
            title="保存导出文件",
            defaultextension=file_extension,
            filetypes=file_types,
            initialname=f"数据导出_{datetime.now().strftime('%Y%m%d_%H%M%S')}{file_extension}",
        )

        if not filename:
            return

        # 关闭对话框
        dialog.destroy()

        # 开始导出
        self._perform_export(
            filename,
            ExportFormat(export_format),
            ExportScope(export_scope),
            export_data,
            selected_columns,
        )

    def _perform_export(
        self,
        filename: str,
        export_format: ExportFormat,
        export_scope: ExportScope,
        data: List[Dict[str, Any]],
        columns: List[Dict[str, Any]],
    ) -> None:
        """执行导出"""
        # 创建进度对象
        self.export_progress = ExportProgress(len(data))

        # 显示进度对话框
        if self.show_progress:
            self._show_progress_dialog()

        # 触发导出开始事件
        if self.on_export_started:
            self.on_export_started(filename, export_scope)

        # 在后台线程中执行导出
        self.current_export_thread = threading.Thread(
            target=self._export_worker,
            args=(filename, export_format, data, columns),
            daemon=True,
        )
        self.current_export_thread.start()

    def _export_worker(
        self,
        filename: str,
        export_format: ExportFormat,
        data: List[Dict[str, Any]],
        columns: List[Dict[str, Any]],
    ) -> None:
        """导出工作线程"""
        success = False
        error_message = ""

        try:
            if export_format == ExportFormat.EXCEL:
                success = self._export_to_excel(filename, data, columns)
            elif export_format == ExportFormat.CSV:
                success = self._export_to_csv(filename, data, columns)
            else:
                error_message = f"不支持的导出格式: {export_format.value}"

        except Exception as e:
            error_message = f"导出失败: {e!s}"
            self.logger.error(f"导出错误: {e}", exc_info=True)

        # 在主线程中处理完成事件
        self.after(
            0, lambda: self._on_export_finished(filename, success, error_message)
        )

    def _export_to_excel(
        self, filename: str, data: List[Dict[str, Any]], columns: List[Dict[str, Any]]
    ) -> bool:
        """导出到Excel文件"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl库未安装,无法导出Excel文件")

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "数据导出"

        # 设置标题样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 写入标题行
        for col_idx, col in enumerate(columns, 1):
            cell = worksheet.cell(
                row=1, column=col_idx, value=col.get("text", col["id"])
            )
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # 写入数据行
        for row_idx, row_data in enumerate(data, 2):
            if self.export_progress and self.export_progress.cancelled:
                return False

            for col_idx, col in enumerate(columns, 1):
                value = row_data.get(col["id"], "")
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border

            # 更新进度
            if self.export_progress:
                self.export_progress.update(
                    row_idx - 1, f"正在导出第 {row_idx - 1} / {len(data)} 行数据..."
                )

                # 触发进度更新事件
                if self.on_export_progress:
                    self.after(0, lambda: self.on_export_progress(self.export_progress))

        # 自动调整列宽
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # 保存文件
        workbook.save(filename)
        return True

    def _export_to_csv(
        self, filename: str, data: List[Dict[str, Any]], columns: List[Dict[str, Any]]
    ) -> bool:
        """导出到CSV文件"""
        with open(filename, "w", newline="", encoding="utf-8-sig") as csvfile:
            writer = csv.writer(csvfile)

            # 写入标题行
            headers = [col.get("text", col["id"]) for col in columns]
            writer.writerow(headers)

            # 写入数据行
            for row_idx, row_data in enumerate(data):
                if self.export_progress and self.export_progress.cancelled:
                    return False

                row = [row_data.get(col["id"], "") for col in columns]
                writer.writerow(row)

                # 更新进度
                if self.export_progress:
                    self.export_progress.update(
                        row_idx + 1, f"正在导出第 {row_idx + 1} / {len(data)} 行数据..."
                    )

                    # 触发进度更新事件
                    if self.on_export_progress:
                        self.after(
                            0, lambda: self.on_export_progress(self.export_progress)
                        )

        return True

    def _show_progress_dialog(self) -> None:
        """显示进度对话框"""
        self.progress_dialog = tk.Toplevel(self)
        self.progress_dialog.title("导出进度")
        self.progress_dialog.geometry("400x150")
        self.progress_dialog.resizable(False, False)
        self.progress_dialog.transient(self)
        self.progress_dialog.grab_set()

        # 居中显示
        self.progress_dialog.update_idletasks()
        x = (self.progress_dialog.winfo_screenwidth() // 2) - (
            self.progress_dialog.winfo_width() // 2
        )
        y = (self.progress_dialog.winfo_screenheight() // 2) - (
            self.progress_dialog.winfo_height() // 2
        )
        self.progress_dialog.geometry(f"+{x}+{y}")

        # 创建进度对话框内容
        main_frame = ttk.Frame(self.progress_dialog, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 进度信息
        self.progress_label = ttk.Label(main_frame, text="准备导出...")
        self.progress_label.pack(pady=(0, 10))

        # 进度条
        self.progress_bar = ttk.Progressbar(main_frame, mode="determinate", maximum=100)
        self.progress_bar.pack(fill=tk.X, pady=(0, 10))

        # 取消按钮
        cancel_btn = ttk.Button(main_frame, text="取消", command=self._cancel_export)
        cancel_btn.pack()

        # 开始更新进度
        self._update_progress_dialog()

    def _update_progress_dialog(self) -> None:
        """更新进度对话框"""
        if not self.progress_dialog or not self.export_progress:
            return

        try:
            # 更新进度条
            self.progress_bar["value"] = self.export_progress.percentage

            # 更新进度文本
            if self.export_progress.message:
                self.progress_label.config(text=self.export_progress.message)

            # 如果导出未完成,继续更新
            if (
                self.current_export_thread
                and self.current_export_thread.is_alive()
                and not self.export_progress.cancelled
            ):
                self.after(100, self._update_progress_dialog)

        except tk.TclError:
            # 对话框已关闭
            pass

    def _cancel_export(self) -> None:
        """取消导出"""
        if self.export_progress:
            self.export_progress.cancel()

        if self.progress_dialog:
            self.progress_dialog.destroy()
            self.progress_dialog = None

        self.logger.info("用户取消了导出操作")

    def _on_export_finished(
        self, filename: str, success: bool, error_message: str = ""
    ) -> None:
        """导出完成处理"""
        # 关闭进度对话框
        if self.progress_dialog:
            self.progress_dialog.destroy()
            self.progress_dialog = None

        # 显示结果消息
        if success:
            messagebox.showinfo("导出成功", f"数据已成功导出到:\n{filename}")
        else:
            messagebox.showerror("导出失败", error_message or "导出过程中发生未知错误")

        # 触发导出完成事件
        if self.on_export_completed:
            self.on_export_completed(filename, success)

        # 清理状态
        self.current_export_thread = None
        self.export_progress = None

        self.logger.info(f"导出完成: {filename}, 成功: {success}")

    def export_data_directly(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        export_format: ExportFormat = ExportFormat.EXCEL,
        columns: Optional[List[Dict[str, Any]]] = None,
    ) -> bool:
        """直接导出数据(不显示对话框)

        Args:
            data: 要导出的数据
            filename: 导出文件名
            export_format: 导出格式
            columns: 要导出的列(None表示导出所有列)

        Returns:
            导出是否成功
        """
        if not data:
            self.logger.warning("没有数据可以导出")
            return False

        export_columns = columns or self.columns

        try:
            if export_format == ExportFormat.EXCEL:
                return self._export_to_excel(filename, data, export_columns)
            if export_format == ExportFormat.CSV:
                return self._export_to_csv(filename, data, export_columns)
            self.logger.error(f"不支持的导出格式: {export_format.value}")
            return False

        except Exception as e:
            self.logger.error(f"直接导出失败: {e}", exc_info=True)
            return False

    def cleanup(self) -> None:
        """清理资源"""
        # 取消正在进行的导出
        if self.export_progress:
            self.export_progress.cancel()

        # 等待导出线程结束
        if self.current_export_thread and self.current_export_thread.is_alive():
            self.current_export_thread.join(timeout=1.0)

        # 关闭进度对话框
        if self.progress_dialog:
            self.progress_dialog.destroy()

        # 清理回调
        self.on_export_started = None
        self.on_export_completed = None
        self.on_export_progress = None

        self.logger.info("导出组件资源已清理")
